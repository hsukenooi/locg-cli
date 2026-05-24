"""Tests for the Excel import / reconciliation pipeline (Unit 2)."""
from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import pytest

FIXTURES = Path(__file__).parent / "fixtures"
SAMPLE_XLSX = FIXTURES / "collection_export_sample.xlsx"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def make_cache(tmp_path: Path):
    from locg.collection_cache import CollectionCache
    return CollectionCache(
        path=tmp_path / "collection.json",
        lock_path=tmp_path / "collection.lock",
        audit_path=tmp_path / "import-history.jsonl",
    )


def make_agent_win_row(
    publisher: str = "Marvel",
    series: str = "Amazing Spider-Man",
    full_title: str = "Amazing Spider-Man #300",
    release_date: str = "1988-05-10",
    needs_manual_series: bool = False,
    needs_manual_variant: bool = False,
    gixen_item_id: str | None = "42",
    pushed: str | None = None,
) -> dict[str, Any]:
    return {
        "publisher_name": publisher,
        "series_name": series,
        "full_title": full_title,
        "release_date": release_date,
        "in_collection": 1,
        "in_wish_list": 0,
        "marked_read": 0,
        "my_rating": None,
        "media_format": "Print",
        "price_paid": None,
        "date_purchased": None,
        "condition": None,
        "notes": None,
        "tags": None,
        "storage_box": None,
        "owner": None,
        "purchase_store": None,
        "signature": None,
        "slabbing": None,
        "grading": None,
        "grading_company": None,
        "local_added_at": "2024-01-01T00:00:00.000000Z",
        "local_added_seq": 1,
        "pushed_to_locg_at": pushed,
        "last_seen_in_export_at": None,
        "source": "agent_win",
        "needs_manual_variant": needs_manual_variant,
        "needs_manual_series_canonical": needs_manual_series,
        "metron_id": None,
        "gixen_item_id": gixen_item_id,
        "previous_full_title": None,
    }


# ---------------------------------------------------------------------------
# parse_xlsx
# ---------------------------------------------------------------------------

def test_parse_xlsx_row_count():
    """parse_xlsx returns rows; count == max_row - 1 (minus header)."""
    from locg.collection_io import parse_xlsx
    rows = parse_xlsx(SAMPLE_XLSX)
    assert len(rows) > 0
    # Sample file has 2353 data rows (2354 including header)
    assert len(rows) == 2353


def test_parse_xlsx_row_shape():
    """Every parsed row has all 21 LOCG column keys."""
    from locg.collection_io import parse_xlsx
    from locg.collection_cache import LOCG_COLUMNS
    rows = parse_xlsx(SAMPLE_XLSX)
    for row in rows[:5]:
        for col in LOCG_COLUMNS:
            assert col in row, f"Missing column: {col}"


def test_parse_xlsx_first_row_values():
    """First data row matches known fixture values."""
    from locg.collection_io import parse_xlsx
    rows = parse_xlsx(SAMPLE_XLSX)
    first = rows[0]
    assert first["publisher_name"] == "Image Comics"
    assert first["series_name"] == "1963 (1993)"
    assert first["full_title"] == "1963 #6"
    assert first["in_collection"] == 1


def test_parse_xlsx_header_mismatch_raises(tmp_path):
    """A file with mismatched column headers raises RuntimeError before any row is read."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Publisher", "Wrong Column"])  # bad headers
    bad_path = tmp_path / "bad.xlsx"
    wb.save(bad_path)

    from locg.collection_io import parse_xlsx
    with pytest.raises(RuntimeError, match="header"):
        parse_xlsx(bad_path)


def test_parse_xlsx_file_too_large_rejected(tmp_path):
    """Files larger than 10 MB are rejected before parsing."""
    big_file = tmp_path / "big.xlsx"
    big_file.write_bytes(b"x" * (10 * 1024 * 1024 + 1))
    from locg.collection_io import parse_xlsx
    with pytest.raises(RuntimeError, match="10 MB"):
        parse_xlsx(big_file)


# ---------------------------------------------------------------------------
# import_xlsx — Phase 2 standard merge (happy paths)
# ---------------------------------------------------------------------------

def test_import_xlsx_populates_empty_cache(tmp_path):
    """Importing into an empty cache inserts all rows with source='locg_export'."""
    from locg.collection_io import import_xlsx
    cache = make_cache(tmp_path)
    result = import_xlsx(SAMPLE_XLSX, cache)
    assert result["added"] > 0
    assert result["updated"] == 0
    payload = cache.load()
    assert len(payload["comics"]) == result["added"]
    assert all(r["source"] == "locg_export" for r in payload["comics"])


def test_import_xlsx_sets_pushed_to_locg_at(tmp_path):
    """Rows imported from LOCG have pushed_to_locg_at set (they're already in LOCG)."""
    from locg.collection_io import import_xlsx
    cache = make_cache(tmp_path)
    import_xlsx(SAMPLE_XLSX, cache)
    payload = cache.load()
    for row in payload["comics"]:
        assert row["pushed_to_locg_at"] is not None


def test_reimport_same_xlsx_unchanged(tmp_path):
    """Re-importing the same xlsx updates last_seen_in_export_at but adds no new rows."""
    from locg.collection_io import import_xlsx
    cache = make_cache(tmp_path)
    r1 = import_xlsx(SAMPLE_XLSX, cache)
    r2 = import_xlsx(SAMPLE_XLSX, cache)
    assert r2["added"] == 0
    assert r2["updated"] >= 0  # last_seen_in_export_at updated


def test_import_updates_last_full_import(tmp_path):
    """After import, last_full_import is set in the payload."""
    from locg.collection_io import import_xlsx
    cache = make_cache(tmp_path)
    import_xlsx(SAMPLE_XLSX, cache)
    payload = cache.load()
    assert payload["last_full_import"] is not None
    assert payload["last_import_source"] == str(SAMPLE_XLSX)


def test_import_builds_series_name_index(tmp_path):
    """After import, series_name_index is non-empty."""
    from locg.collection_io import import_xlsx
    cache = make_cache(tmp_path)
    import_xlsx(SAMPLE_XLSX, cache)
    payload = cache.load()
    assert len(payload["series_name_index"]) > 0


# ---------------------------------------------------------------------------
# import_xlsx — Phase 2: cache-only rows preserved
# ---------------------------------------------------------------------------

def test_cache_only_agent_win_survives_import(tmp_path):
    """agent_win rows not in the xlsx are preserved, not deleted (v1 preserves)."""
    from locg.collection_cache import CollectionCache
    from locg.collection_io import import_xlsx

    cache = make_cache(tmp_path)
    win_row = make_agent_win_row(
        publisher="DC",
        series="Batman (1940 - 2011)",
        full_title="Batman #999 (Not in Export)",
        release_date="2999-01-01",
    )

    def add_win(payload):
        payload["comics"].append(win_row)

    cache.apply(add_win, command="pre-import")
    import_xlsx(SAMPLE_XLSX, cache)

    payload = cache.load()
    titles = {r["full_title"] for r in payload["comics"]}
    assert "Batman #999 (Not in Export)" in titles


def test_pushed_not_in_export_possibly_removed_logged(tmp_path):
    """A pushed row not appearing in the re-export logs a 'possibly_removed' audit record."""
    from locg.collection_cache import CollectionCache
    from locg.collection_io import import_xlsx

    cache = make_cache(tmp_path)
    # Add a "pushed" agent_win row that won't appear in the xlsx
    ghost_row = make_agent_win_row(
        publisher="Nonexistent",
        series="Ghost Series (2099)",
        full_title="Ghost #1",
        release_date="2099-01-01",
        pushed="2024-01-01T00:00:00Z",
    )

    def add_ghost(payload):
        payload["comics"].append(ghost_row)

    cache.apply(add_ghost, command="pre-import")
    result = import_xlsx(SAMPLE_XLSX, cache)

    # The row must still be in the cache
    payload = cache.load()
    titles = {r["full_title"] for r in payload["comics"]}
    assert "Ghost #1" in titles

    # Audit log must have a possibly_removed record
    audit_lines = (tmp_path / "import-history.jsonl").read_text().strip().splitlines()
    audit_types = [json.loads(l)["type"] for l in audit_lines]
    assert "possibly_removed" in audit_types


# ---------------------------------------------------------------------------
# import_xlsx — Phase 1 reconciliation (R60)
# ---------------------------------------------------------------------------

def test_reconciliation_best_guess_row_resolved(tmp_path):
    """A best-guess agent_win row matched by reconciliation heuristic gets identity rewritten."""
    from locg.collection_cache import CollectionCache
    from locg.collection_io import import_xlsx

    cache = make_cache(tmp_path)
    # Add a row that matches what's in the xlsx — use a known series/issue from fixture
    # The fixture has "1963 #6" by Image Comics
    unresolved = make_agent_win_row(
        publisher="Image Comics",
        series="1963",  # bare name, no year — needs reconciliation
        full_title="1963 #6",
        release_date="1993-11-08",
        needs_manual_series=True,
    )
    unresolved["local_added_at"] = "2024-01-02T00:00:00.000000Z"
    unresolved["local_added_seq"] = 1

    def add_unresolved(payload):
        payload["comics"].append(unresolved)

    cache.apply(add_unresolved, command="pre-import")
    import_xlsx(SAMPLE_XLSX, cache)

    payload = cache.load()
    # Find the reconciled row
    matched = [r for r in payload["comics"] if r["full_title"] == "1963 #6"
               and r.get("gixen_item_id") == "42"]
    assert len(matched) == 1, "Reconciled row not found (expected exactly one)"
    assert matched[0]["needs_manual_series_canonical"] is False
    assert matched[0]["series_name"] == "1963 (1993)"  # LOCG canonical name
    # Tracking fields preserved
    assert matched[0]["gixen_item_id"] == "42"
    assert matched[0]["local_added_at"] == "2024-01-02T00:00:00.000000Z"


def test_reconciliation_vol_mismatch_not_reconciled(tmp_path):
    """A row with mismatching (Vol. N) annotation is NOT reconciled per R60."""
    from locg.collection_cache import CollectionCache
    from locg.collection_io import import_xlsx

    cache = make_cache(tmp_path)
    # Force a Vol mismatch: cache says "Amazing Spider-Man (Vol. 2)" but xlsx has Vol. 1
    bad_match = make_agent_win_row(
        publisher="Marvel Comics",
        series="Amazing Spider-Man (Vol. 2)",  # wrong vol
        full_title="Amazing Spider-Man #300",
        release_date="1988-05-10",
        needs_manual_series=True,
    )

    def add_row(payload):
        payload["comics"].append(bad_match)

    cache.apply(add_row, command="pre-import")
    import_xlsx(SAMPLE_XLSX, cache)

    payload = cache.load()
    remaining_manual = [
        r for r in payload["comics"]
        if r.get("gixen_item_id") == "42"
        and r.get("needs_manual_series_canonical") is True
    ]
    assert len(remaining_manual) == 1, "Row should still be flagged needs_manual_series_canonical"


# ---------------------------------------------------------------------------
# import_xlsx — renamed full_title persistence (R67)
# ---------------------------------------------------------------------------

def test_renamed_full_title_persists_previous(tmp_path):
    """When LOCG renames a full_title, previous_full_title is set for one cycle."""
    from locg.collection_cache import CollectionCache
    from locg.collection_io import import_xlsx

    cache = make_cache(tmp_path)
    # Seed cache with a row using the "old" title
    old_title = "1963 #6 (Old Name)"
    old_row = {
        "publisher_name": "Image Comics",
        "series_name": "1963 (1993)",
        "full_title": old_title,
        "release_date": "1993-11-08",
        "in_collection": 1,
        "in_wish_list": 0,
        "marked_read": 0,
        "my_rating": None,
        "media_format": "Print",
        "price_paid": None,
        "date_purchased": None,
        "condition": None,
        "notes": None,
        "tags": None,
        "storage_box": None,
        "owner": None,
        "purchase_store": None,
        "signature": None,
        "slabbing": None,
        "grading": None,
        "grading_company": None,
        "local_added_at": "2024-01-01T00:00:00.000000Z",
        "local_added_seq": 1,
        "pushed_to_locg_at": "2024-01-01T00:00:00Z",
        "last_seen_in_export_at": None,
        "source": "locg_export",
        "needs_manual_variant": False,
        "needs_manual_series_canonical": False,
        "metron_id": None,
        "gixen_item_id": None,
        "previous_full_title": None,
    }

    def seed(payload):
        payload["comics"].append(old_row)

    cache.apply(seed, command="seed")
    import_xlsx(SAMPLE_XLSX, cache)

    payload = cache.load()
    # The import should have matched by (publisher, series, release_date) and updated title
    renamed = [r for r in payload["comics"]
               if r.get("previous_full_title") == old_title]
    assert len(renamed) == 1, "Expected exactly one row with previous_full_title set"
    assert renamed[0]["full_title"] == "1963 #6"  # LOCG's canonical title


# ---------------------------------------------------------------------------
# import_xlsx — behavioral drift detection (F5)
# ---------------------------------------------------------------------------

def test_behavioral_drift_detected(tmp_path):
    """A changed user-managed column logs a behavioral_drift audit record."""
    from locg.collection_cache import CollectionCache
    from locg.collection_io import import_xlsx

    cache = make_cache(tmp_path)
    # Seed with a row that has a non-null my_rating that the import will clear
    row_with_rating = {
        "publisher_name": "Image Comics",
        "series_name": "1963 (1993)",
        "full_title": "1963 #6",
        "release_date": "1993-11-08",
        "in_collection": 1,
        "in_wish_list": 0,
        "marked_read": 0,
        "my_rating": 9.0,  # user set this; xlsx returns None → drift
        "media_format": "Print",
        "price_paid": None,
        "date_purchased": None,
        "condition": None,
        "notes": None,
        "tags": None,
        "storage_box": None,
        "owner": None,
        "purchase_store": None,
        "signature": None,
        "slabbing": None,
        "grading": None,
        "grading_company": None,
        "local_added_at": "2024-01-01T00:00:00.000000Z",
        "local_added_seq": 1,
        "pushed_to_locg_at": "2024-01-01T00:00:00Z",
        "last_seen_in_export_at": None,
        "source": "locg_export",
        "needs_manual_variant": False,
        "needs_manual_series_canonical": False,
        "metron_id": None,
        "gixen_item_id": None,
        "previous_full_title": None,
    }

    def seed(payload):
        payload["comics"].append(row_with_rating)

    cache.apply(seed, command="seed")
    import_xlsx(SAMPLE_XLSX, cache)

    audit_lines = (tmp_path / "import-history.jsonl").read_text().strip().splitlines()
    audit_types = [json.loads(l)["type"] for l in audit_lines]
    assert "behavioral_drift" in audit_types


# ---------------------------------------------------------------------------
# import_xlsx — series_name_index rebuilt from locg_export only
# ---------------------------------------------------------------------------

def test_series_name_index_after_import(tmp_path):
    """series_name_index is rebuilt from locg_export rows only after import."""
    from locg.collection_cache import CollectionCache
    from locg.collection_io import import_xlsx

    cache = make_cache(tmp_path)
    # Add an agent_win row — its series must NOT appear in the index
    win = make_agent_win_row(series="Totally Made Up Series (2099)")

    def add_win(payload):
        payload["comics"].append(win)

    cache.apply(add_win, command="pre-import")
    import_xlsx(SAMPLE_XLSX, cache)

    payload = cache.load()
    index = payload["series_name_index"]
    assert not any("totally made up" in k.lower() for k in index)
    # But fixture series should be present
    assert any("1963" in k for k in index)


# ---------------------------------------------------------------------------
# import_xlsx — crash recovery: migration_in_progress stays False
# ---------------------------------------------------------------------------

def test_import_never_leaves_migration_flag(tmp_path):
    """After a successful import, migration_in_progress must be False."""
    from locg.collection_io import import_xlsx
    cache = make_cache(tmp_path)
    import_xlsx(SAMPLE_XLSX, cache)
    payload = cache.load()
    assert payload["migration_in_progress"] is False


# ---------------------------------------------------------------------------
# import_xlsx — error paths
# ---------------------------------------------------------------------------

def test_import_nonexistent_file_raises(tmp_path):
    from locg.collection_io import import_xlsx
    cache = make_cache(tmp_path)
    with pytest.raises((RuntimeError, FileNotFoundError, OSError)):
        import_xlsx(tmp_path / "does_not_exist.xlsx", cache)


def test_import_bad_header_raises_before_merge(tmp_path):
    """Header mismatch must raise before any cache mutation."""
    import openpyxl
    from locg.collection_io import import_xlsx

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Wrong", "Headers", "Here"])
    bad_path = tmp_path / "bad_headers.xlsx"
    wb.save(bad_path)

    cache = make_cache(tmp_path)
    with pytest.raises(RuntimeError, match="header"):
        import_xlsx(bad_path, cache)

    # Cache should be untouched — no file created
    assert not (tmp_path / "collection.json").exists()


# ---------------------------------------------------------------------------
# Unit 3: generate_csv
# ---------------------------------------------------------------------------

RECIPE_CSV = FIXTURES / "locg_import_test_recipe.csv"


def _make_ready_row(
    publisher: str = "Marvel Comics",
    series: str = "The Amazing Spider-Man (Vol. 1) (1962 - 1998)",
    full_title: str = "The Amazing Spider-Man #84",
    release_date: str = "1970-05-01",
    price_paid: Any = 27.86,
    date_purchased: Any = "2026-05-22",
) -> dict[str, Any]:
    """Build a minimal agent_win row suitable for CSV export."""
    return {
        "publisher_name": publisher,
        "series_name": series,
        "full_title": full_title,
        "release_date": release_date,
        "in_collection": 1,
        "in_wish_list": 0,
        "marked_read": 0,
        "my_rating": None,
        "media_format": "Print",
        "price_paid": price_paid,
        "date_purchased": date_purchased,
        "condition": None,
        "notes": None,
        "tags": None,
        "storage_box": None,
        "owner": None,
        "purchase_store": "eBay",
        "signature": 0,
        "slabbing": 0,
        "grading": None,
        "grading_company": None,
        "local_added_at": "2026-05-22T10:00:00.000000Z",
        "local_added_seq": 1,
        "pushed_to_locg_at": None,
        "last_seen_in_export_at": None,
        "source": "agent_win",
        "needs_manual_variant": False,
        "needs_manual_series_canonical": False,
        "metron_id": None,
        "gixen_item_id": "12345",
        "previous_full_title": None,
    }


def test_generate_csv_row_count(tmp_path):
    """10 ready rows produce a 10-row CSV (plus header)."""
    import csv
    from locg.collection_io import generate_csv
    rows = [_make_ready_row(full_title=f"ASM #{i}") for i in range(10)]
    out = tmp_path / "out.csv"
    generate_csv(rows, out)
    with open(out, newline="") as f:
        reader = list(csv.reader(f))
    assert len(reader) == 11  # 1 header + 10 data rows


def test_generate_csv_header_order(tmp_path):
    """CSV header matches the canonical 21-column LOCG order."""
    import csv
    from locg.collection_io import LOCG_XLSX_HEADERS, generate_csv
    generate_csv([_make_ready_row()], tmp_path / "out.csv")
    with open(tmp_path / "out.csv", newline="") as f:
        header = next(csv.reader(f))
    assert tuple(header) == LOCG_XLSX_HEADERS


def test_generate_csv_my_rating_blank_in_body(tmp_path):
    """My Rating column is present in header AND body as an empty string (R27)."""
    from locg.collection_io import generate_csv
    generate_csv([_make_ready_row()], tmp_path / "out.csv")
    raw = (tmp_path / "out.csv").read_text()
    lines = raw.splitlines()
    # Header must contain My Rating
    assert "My Rating" in lines[0]
    # Body row: the My Rating field position must be empty
    import csv, io
    reader = list(csv.reader(io.StringIO(raw)))
    header = reader[0]
    my_rating_idx = header.index("My Rating")
    body_row = reader[1]
    assert body_row[my_rating_idx] == ""


def test_generate_csv_empty_queue_header_only(tmp_path):
    """Zero ready rows produces a CSV with only the header line."""
    import csv
    from locg.collection_io import generate_csv
    generate_csv([], tmp_path / "out.csv")
    with open(tmp_path / "out.csv", newline="") as f:
        rows = list(csv.reader(f))
    assert len(rows) == 1  # header only


def test_generate_csv_price_format(tmp_path):
    """Price Paid is formatted as NN.NN with no currency suffix."""
    import csv
    from locg.collection_io import generate_csv
    generate_csv([_make_ready_row(price_paid=27.8600001)], tmp_path / "out.csv")
    with open(tmp_path / "out.csv", newline="") as f:
        rows = list(csv.reader(f))
    header = rows[0]
    price_idx = header.index("Price Paid")
    assert rows[1][price_idx] == "27.86"


def test_generate_csv_negative_price_defaults_to_zero(tmp_path):
    """Negative price_paid defaults to 0.00 (R29)."""
    import csv
    from locg.collection_io import generate_csv
    generate_csv([_make_ready_row(price_paid=-5.0)], tmp_path / "out.csv")
    with open(tmp_path / "out.csv", newline="") as f:
        rows = list(csv.reader(f))
    price_idx = rows[0].index("Price Paid")
    assert rows[1][price_idx] == "0.00"


def test_generate_csv_missing_price_defaults_to_zero(tmp_path):
    """None price_paid defaults to 0.00."""
    import csv
    from locg.collection_io import generate_csv
    generate_csv([_make_ready_row(price_paid=None)], tmp_path / "out.csv")
    with open(tmp_path / "out.csv", newline="") as f:
        rows = list(csv.reader(f))
    price_idx = rows[0].index("Price Paid")
    assert rows[1][price_idx] == "0.00"


def test_generate_csv_date_iso_format(tmp_path):
    """Date Purchased is output as ISO date (YYYY-MM-DD) (R30)."""
    import csv
    from locg.collection_io import generate_csv
    generate_csv([_make_ready_row(date_purchased="2026-05-22")], tmp_path / "out.csv")
    with open(tmp_path / "out.csv", newline="") as f:
        rows = list(csv.reader(f))
    date_idx = rows[0].index("Date Purchased")
    assert rows[1][date_idx] == "2026-05-22"


def test_generate_csv_fixed_fields(tmp_path):
    """In Collection=1, In Wish List=0, Marked Read=0, Media Format=Print, Purchase Store=eBay."""
    import csv
    from locg.collection_io import generate_csv
    generate_csv([_make_ready_row()], tmp_path / "out.csv")
    with open(tmp_path / "out.csv", newline="") as f:
        rows = list(csv.reader(f))
    h, d = rows[0], rows[1]
    assert d[h.index("In Collection")] == "1"
    assert d[h.index("In Wish List")] == "0"
    assert d[h.index("Marked Read")] == "0"
    assert d[h.index("Media Format")] == "Print"
    assert d[h.index("Purchase Store")] == "eBay"
    assert d[h.index("Signature")] == "0"
    assert d[h.index("Slabbing")] == "0"


def test_generate_csv_bitforbit_recipe(tmp_path):
    """CSV output for the validated golden fixture rows matches the recipe bit-for-bit."""
    import csv as _csv
    from locg.collection_io import generate_csv

    # Build the same rows as in locg_import_test_recipe.csv
    with open(RECIPE_CSV, newline="") as f:
        reader = _csv.DictReader(f)
        recipe_rows = list(reader)

    # Reconstruct cache rows from the recipe CSV
    cache_rows = []
    for r in recipe_rows:
        cache_rows.append({
            "publisher_name": r["Publisher Name"],
            "series_name": r["Series Name"],
            "full_title": r["Full Title"],
            "release_date": r["Release Date"],
            "price_paid": float(r["Price Paid"]) if r["Price Paid"] else None,
            "date_purchased": r["Date Purchased"] or None,
            "needs_manual_variant": False,
            "needs_manual_series_canonical": False,
            # All other fields not needed for CSV output
        })

    out = tmp_path / "test_out.csv"
    generate_csv(cache_rows, out)

    generated = out.read_text()
    expected = RECIPE_CSV.read_text()
    assert generated == expected


# ---------------------------------------------------------------------------
# Unit 3: generate_notes_md
# ---------------------------------------------------------------------------

def test_notes_md_ready_count(tmp_path):
    """notes.md correctly counts ready rows."""
    from locg.collection_io import generate_notes_md
    ready = [_make_ready_row() for _ in range(5)]
    out = tmp_path / "out.notes.md"
    generate_notes_md(ready, [], [], out)
    text = out.read_text()
    assert "Ready to upload (5 rows)" in text


def test_notes_md_empty_queue(tmp_path):
    """Zero pending rows produces notes.md noting empty queue."""
    from locg.collection_io import generate_notes_md
    out = tmp_path / "out.notes.md"
    generate_notes_md([], [], [], out)
    text = out.read_text()
    assert "Ready to upload (0 rows)" in text


def test_notes_md_manual_variant_section(tmp_path):
    """Variant rows appear in the variants section, not ready section."""
    from locg.collection_io import generate_notes_md
    variant_row = _make_ready_row(full_title="ASM #300 Newsstand")
    variant_row["needs_manual_variant"] = True
    out = tmp_path / "out.notes.md"
    generate_notes_md([], [variant_row], [], out)
    text = out.read_text()
    assert "Needs manual handling — variants (1 rows)" in text
    assert "ASM #300 Newsstand" in text


def test_notes_md_manual_series_section(tmp_path):
    """Series-canonical rows appear in the series canonical section."""
    from locg.collection_io import generate_notes_md
    series_row = _make_ready_row(series="Unknown Series")
    series_row["needs_manual_series_canonical"] = True
    out = tmp_path / "out.notes.md"
    generate_notes_md([], [], [series_row], out)
    text = out.read_text()
    assert "Needs manual handling — series canonical (1 rows)" in text
    assert "Unknown Series" in text


# ---------------------------------------------------------------------------
# Unit 3: _pending_push_rows
# ---------------------------------------------------------------------------

def test_pending_push_rows_partitions(tmp_path):
    """_pending_push_rows correctly partitions ready / manual_variant / manual_series."""
    from locg.collection_io import _pending_push_rows

    r = _make_ready_row()
    v = _make_ready_row(full_title="ASM #300 Newsstand")
    v["needs_manual_variant"] = True
    s = _make_ready_row(series="Unknown")
    s["needs_manual_series_canonical"] = True
    already_pushed = _make_ready_row(full_title="Pushed #1")
    already_pushed["pushed_to_locg_at"] = "2030-01-01T00:00:00Z"  # future; not pending
    already_pushed["local_added_at"] = "2026-01-01T00:00:00Z"

    payload = {"comics": [r, v, s, already_pushed]}
    ready, mv, ms = _pending_push_rows(payload)
    assert len(ready) == 1
    assert len(mv) == 1
    assert len(ms) == 1
    assert ready[0]["full_title"] == _make_ready_row()["full_title"]


def test_pending_push_already_pushed_excluded(tmp_path):
    """Rows where local_added_at <= pushed_to_locg_at are not pending."""
    from locg.collection_io import _pending_push_rows

    row = _make_ready_row()
    row["pushed_to_locg_at"] = "2030-01-01T00:00:00Z"  # pushed far in the future
    row["local_added_at"] = "2026-01-01T00:00:00Z"  # added before push timestamp

    ready, mv, ms = _pending_push_rows({"comics": [row]})
    assert len(ready) == 0
    assert len(mv) == 0
    assert len(ms) == 0


# ---------------------------------------------------------------------------
# import_xlsx — wish-list cache
# ---------------------------------------------------------------------------

def test_import_xlsx_writes_wish_list_cache(tmp_path, monkeypatch):
    """import_xlsx writes wish-list.json alongside collection.json."""
    from locg.collection_io import import_xlsx, wish_list_cache_path
    cache = make_cache(tmp_path)
    import_xlsx(SAMPLE_XLSX, cache)
    wl_path = wish_list_cache_path()
    assert wl_path.exists()
    data = json.loads(wl_path.read_text())
    assert "updated_at" in data
    assert "items" in data


def test_wish_list_cache_item_count(tmp_path):
    """wish-list.json contains exactly the rows where in_wish_list == 1."""
    from locg.collection_io import import_xlsx, wish_list_cache_path
    cache = make_cache(tmp_path)
    import_xlsx(SAMPLE_XLSX, cache)
    payload = cache.load()
    expected = sum(1 for r in payload["comics"] if r.get("in_wish_list") == 1)
    data = json.loads(wish_list_cache_path().read_text())
    assert len(data["items"]) == expected
    assert expected > 0  # fixture has wish-list rows


def test_wish_list_cache_item_shape(tmp_path):
    """Each wish-list cache entry has name (from full_title) and id: null."""
    from locg.collection_io import import_xlsx, wish_list_cache_path
    cache = make_cache(tmp_path)
    import_xlsx(SAMPLE_XLSX, cache)
    payload = cache.load()
    xl_wish = {r["full_title"] for r in payload["comics"] if r.get("in_wish_list") == 1}
    data = json.loads(wish_list_cache_path().read_text())
    for item in data["items"]:
        assert item["name"] in xl_wish
        assert item["id"] is None


def test_wish_list_cache_updated_at_is_iso_timestamp(tmp_path):
    """wish-list.json envelope contains a valid ISO 8601 UTC timestamp."""
    from datetime import datetime, timezone
    from locg.collection_io import import_xlsx, wish_list_cache_path
    cache = make_cache(tmp_path)
    import_xlsx(SAMPLE_XLSX, cache)
    data = json.loads(wish_list_cache_path().read_text())
    ts = data["updated_at"]
    # Should parse without error and be timezone-aware
    parsed = datetime.fromisoformat(ts)
    assert parsed.tzinfo is not None


def test_wish_list_cache_overwritten_on_reimport(tmp_path):
    """Re-importing overwrites wish-list.json rather than appending."""
    from locg.collection_io import import_xlsx, wish_list_cache_path
    cache = make_cache(tmp_path)
    import_xlsx(SAMPLE_XLSX, cache)
    first_mtime = wish_list_cache_path().stat().st_mtime_ns
    import_xlsx(SAMPLE_XLSX, cache)
    second_mtime = wish_list_cache_path().stat().st_mtime_ns
    # Second import must have produced a fresh file
    assert second_mtime >= first_mtime
    data = json.loads(wish_list_cache_path().read_text())
    assert isinstance(data["items"], list)


def test_wish_list_cache_empty_when_no_wish_list_rows(tmp_path, monkeypatch):
    """An XLSX with zero wish-list rows writes items: [] without error."""
    import openpyxl
    from locg.collection_io import import_xlsx, wish_list_cache_path, LOCG_XLSX_HEADERS

    # Build a minimal XLSX with one collection-only row (in_wish_list=0)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(list(LOCG_XLSX_HEADERS))
    ws.append([
        "Marvel", "X-Men", "X-Men #1", "1963-09-01",
        1, 0, 0, None, "Print", None, None, None, None,
        None, None, None, None, None, None, None, None,
    ])
    xlsx_path = tmp_path / "no_wish.xlsx"
    wb.save(xlsx_path)

    cache = make_cache(tmp_path)
    import_xlsx(xlsx_path, cache)
    data = json.loads(wish_list_cache_path().read_text())
    assert data["items"] == []
    assert "updated_at" in data
